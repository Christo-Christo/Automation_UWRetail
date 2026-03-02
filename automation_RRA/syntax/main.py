import os
import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

from syntax.RRA import read_data, process_final_data, generate_all_tables


# =========================================================
# Helper: nama output mengikuti DATA RAW
# =========================================================
def build_output_name_from_raw(raw_data_path, suffix):
    name, ext = os.path.splitext(os.path.basename(raw_data_path))
    return f"{name}_{suffix}{ext}"


# =========================================================
# Helper: formatting Excel (HEADER + GRAND TOTAL + WIDTH)
# =========================================================
def format_excel_file(file_path):
    wb = load_workbook(file_path)

    header_fill = PatternFill(
        start_color="00B0F0",
        end_color="00B0F0",
        fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF")

    for ws in wb.worksheets:

        # ---------- HEADER ----------
        if ws.max_row >= 1:
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font

        # ---------- GRAND TOTAL ----------
        for row in ws.iter_rows(min_row=2):
            if str(row[0].value).strip().lower() == "grand total":
                for cell in row:
                    cell.fill = header_fill
                    cell.font = header_font

        # ---------- AUTO COLUMN WIDTH ----------
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter

            for cell in col:
                try:
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    wb.save(file_path)


# =========================================================
# FILE 1 — DATA + PREVIOUS
# =========================================================
def write_data_file(final, previous, output_path, raw_data_path):
    print("📝 Writing DATA file...")

    output_file = os.path.join(
        output_path,
        build_output_name_from_raw(raw_data_path, "processed")
    )

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        final.to_excel(writer, sheet_name="Data", index=False)
        previous.to_excel(writer, sheet_name="Previous", index=False)

    format_excel_file(output_file)
    print(f"✅ Data file saved: {output_file}")


# =========================================================
# FILE 2 — RC
# =========================================================
def write_rc_file(rc, rcsa_only, rc_sheets, output_path, raw_data_path):
    print("📝 Writing RC file...")

    output_file = os.path.join(
        output_path,
        build_output_name_from_raw(raw_data_path, "RC")
    )

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        rc.to_excel(writer, sheet_name="RC", index=False)
        rcsa_only.to_excel(writer, sheet_name="RCSA", index=False)

        for sheet_name, df in rc_sheets.items():
            df.to_excel(writer, sheet_name=sheet_name[:31], index=False)

    format_excel_file(output_file)
    print(f"✅ RC file saved: {output_file}")


# =========================================================
# FILE 3 — RB
# =========================================================
def write_rb_file(rb, rb_sheets, output_path, raw_data_path):
    print("📝 Writing RB file...")

    output_file = os.path.join(
        output_path,
        build_output_name_from_raw(raw_data_path, "RB")
    )

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        rb.to_excel(writer, sheet_name="RB", index=False)

        for sheet_name, df in rb_sheets.items():
            df.to_excel(writer, sheet_name=sheet_name[:31], index=False)

    format_excel_file(output_file)
    print(f"✅ RB file saved: {output_file}")


# =========================================================
# MAIN (LOGIKA ASLI — TIDAK DIUBAH)
# =========================================================
def main(input_path):

    print("📥 Reading input parameters...")

    input_df = pd.read_excel(
        input_path,
        sheet_name="Input",
        engine="openpyxl",
        header=None
    )

    params = {}
    for _, row in input_df.iterrows():
        key = str(row[0]).strip().replace(":", "")
        params[key] = row[1]

    input_year = int(params["Year"])
    file_path_data = params["File Path Data"]
    file_path_output = params["File Path Output"]

    print("📊 Reading raw data & reference sheets...")
    data, previous, RCSA = read_data(file_path_data, input_path)

    print("🔄 Processing final data...")
    final, col_name = process_final_data(data, previous, input_year)

    print("⚙️ Generating RC & RB tables (this step may take time)...")
    results = generate_all_tables(final, RCSA, input_year, col_name)

    print("📦 Preparing output datasets...")

    final_data = results["Data"]
    rc = results["RC"]
    rb = results["RB"]

    rc_sheets = {
        k: v for k, v in results.items()
        if "RCSA" in k or "NON-RCSA" in k
    }

    rb_sheets = {
        k: v for k, v in results.items()
        if k not in rc_sheets and k not in ["Data", "RC", "RB"]
    }

    rcsa_only = rc[rc["RCSA/NON-RCSA"] == "RCSA"]

    write_data_file(final_data, previous, file_path_output, file_path_data)
    write_rc_file(rc, rcsa_only, rc_sheets, file_path_output, file_path_data)
    write_rb_file(rb, rb_sheets, file_path_output, file_path_data)

    print("🎉 All processes completed successfully!")